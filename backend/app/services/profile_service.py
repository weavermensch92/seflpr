import json
import logging
import uuid
from collections import Counter
from datetime import date
from typing import Optional
from fastapi import HTTPException, UploadFile, status
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.repositories.profile_repo import ProfileRepository
from app.schemas.profile import (
    ProfileCreate, ProfileUpdate, ProfileResponse,
    FreeTextParseRequest, FreeTextParseResponse, ParsedProfileItem, BulkConfirmRequest,
    IngestResponse, AISummary, TimelineEntry, DashboardResponse, TagCount,
)
from app.models.profile import PersonalProfile, ProfileSource, ProfileType

logger = logging.getLogger(__name__)


def _parse_date(date_str: Optional[str]) -> Optional[date]:
    """YYYY-MM 또는 YYYY-MM-DD 형식의 문자열을 date 객체로 변환."""
    if not date_str or not isinstance(date_str, str):
        return None
    try:
        parts = date_str.split("-")
        year = int(parts[0])
        month = int(parts[1]) if len(parts) > 1 else 1
        day = int(parts[2]) if len(parts) > 2 else 1
        return date(year, month, day)
    except (ValueError, IndexError):
        return None


def _to_response(profile) -> ProfileResponse:
    return ProfileResponse(
        id=str(profile.id),
        profile_type=profile.profile_type,
        title=profile.title,
        organization=profile.organization,
        role=profile.role,
        description=getattr(profile, "_description_plain", None),
        start_date=profile.start_date,
        end_date=profile.end_date,
        tags=profile.tags,
        metadata=profile.metadata_,
        is_ai_memory=profile.is_ai_memory,
        ai_interpreted_content=getattr(profile, "_ai_interpreted_content_plain", None),
        enrichment_status=getattr(profile, "enrichment_status", "none"),
        ai_summary_json=getattr(profile, "ai_summary_json", None),
        sort_order=profile.sort_order,
        source=profile.source,
    )


class ProfileService:
    def __init__(self, db: AsyncSession):
        self.db = db
        self.repo = ProfileRepository(db)

    async def list_profiles(self, user_id: str) -> list[ProfileResponse]:
        profiles = await self.repo.list_by_user(user_id)
        return [_to_response(p) for p in profiles]

    async def get_profile(self, profile_id: str, user_id: str) -> ProfileResponse:
        p = await self.repo.get(profile_id, user_id)
        if not p:
            raise HTTPException(status_code=404, detail="프로필을 찾을 수 없습니다.")
        return _to_response(p)

    async def create_profile(self, user_id: str, data: ProfileCreate) -> ProfileResponse:
        raw = data.model_dump(by_alias=False)
        raw["metadata_"] = raw.pop("metadata_", None)
        p = await self.repo.create(user_id, raw)
        return _to_response(p)

    async def update_profile(self, profile_id: str, user_id: str, data: ProfileUpdate) -> ProfileResponse:
        p = await self.repo.get(profile_id, user_id)
        if not p:
            raise HTTPException(status_code=404, detail="프로필을 찾을 수 없습니다.")
        raw = {k: v for k, v in data.model_dump(by_alias=False, exclude_none=True).items()}
        if "metadata_" in raw:
            raw["metadata_"] = raw.pop("metadata_")
        updated = await self.repo.update(p, raw)
        return _to_response(updated)

    async def delete_profile(self, profile_id: str, user_id: str) -> None:
        p = await self.repo.get(profile_id, user_id)
        if not p:
            raise HTTPException(status_code=404, detail="프로필을 찾을 수 없습니다.")
        await self.repo.delete(p)

    async def parse_free_text(self, user_id: str, req: FreeTextParseRequest) -> FreeTextParseResponse:
        """자유 텍스트를 GPT-4o-mini로 파싱하여 프로필 항목 목록 반환."""
        from app.models.user import User
        from sqlalchemy import select
        
        u_id = uuid.UUID(user_id)
        query = select(User).where(User.id == u_id)
        result = await self.db.execute(query)
        user = result.scalar_one_or_none()
        
        if not user:
             raise HTTPException(status_code=404, detail="사용자를 찾을 수 없습니다.")

        # 어드민은 포인트 체크 생략
        if not user.is_admin:
            if user.point_balance <= 0:
                raise HTTPException(
                    status_code=status.HTTP_402_PAYMENT_REQUIRED,
                    detail="AI 자동 분류는 포인트 충전 후 사용 가능합니다."
                )

        from app.core.config import settings

        if not settings.OPENAI_API_KEY:
            # API 키 없을 때 목업 응답
            return FreeTextParseResponse(items=[
                ParsedProfileItem(
                    profile_type="free_text",
                    title="자유 텍스트 입력",
                    description=req.text,
                )
            ])

        from openai import AsyncOpenAI
        client = AsyncOpenAI(api_key=settings.OPENAI_API_KEY)

        prompt = f"""아래 텍스트에서 이력/경력 관련 정보를 추출하여 JSON 객체로 반환하세요.
반드시 "items"라는 키 아래에 배열을 담아서 반환해야 합니다.

각 항목은 다음 필드를 포함합니다:
- profile_type: education|work_experience|project|certification|skill|activity|award|strength|weakness|motivation|free_text
- title: 항목 제목 (필수)
- organization: 학교/회사/단체명
- role: 역할/직책
- description: 상세 내용
- start_date: YYYY-MM 형식
- end_date: YYYY-MM 형식
- tags: 관련 키워드 배열
- metadata: 추가 정보 (education이면 {{"major": "...", "gpa": "...", "gpa_scale": "..."}}, project면 {{"tech_stack": "...", "team_size": "...", "outcome": "..."}})

텍스트:
{req.text}

JSON 형식의 객체만 반환하세요.
{{ "items": [{{...}}, {{...}}] }}
"""

        resp = await client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": prompt}],
            response_format={"type": "json_object"},
            temperature=0.2,
        )
        raw = resp.choices[0].message.content
        data = json.loads(raw)
        items_raw = data.get("items", [])
        
        # 낱개 항목으로 변환
        items = [ParsedProfileItem(**item) for item in items_raw]
        return FreeTextParseResponse(items=items)

    async def parse_url(self, user_id: str, url: str) -> FreeTextParseResponse:
        """외부 URL의 텍스트를 추출하여 AI로 파싱."""
        import httpx
        from bs4 import BeautifulSoup

        try:
            async with httpx.AsyncClient(timeout=10.0, follow_redirects=True) as client:
                resp = await client.get(url, headers={
                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
                })
                
                if resp.status_code in (401, 403):
                    raise HTTPException(
                        status_code=status.HTTP_403_FORBIDDEN,
                        detail="해당 링크에 대한 접근 권한이 없어 내용을 가져올 수 없습니다."
                    )
                resp.raise_for_status()
                
                soup = BeautifulSoup(resp.text, "html.parser")
                # 스크립트, 스타일 제거
                for s in soup(["script", "style"]):
                    s.decompose()
                
                text = soup.get_text(separator="\n", strip=True)
                # 너무 긴 경우 절약
                text = text[:8000]
                
                return await self.parse_free_text(user_id, FreeTextParseRequest(text=text))
                
        except httpx.HTTPStatusError as e:
            if e.response.status_code in (401, 403):
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="접근 권한이 없어서 내용을 반영할 수 없습니다."
                )
            raise HTTPException(status_code=e.response.status_code, detail=f"페이지를 불러오지 못했습니다: {e.response.reason_phrase}")
        except Exception as e:
            if isinstance(e, HTTPException): raise e
            raise HTTPException(status_code=500, detail=f"링크 분석 중 오류 발생: {str(e)}")

    async def bulk_confirm(self, user_id: str, req: BulkConfirmRequest) -> list[ProfileResponse]:
        """파싱 결과를 확인 후 일괄 저장."""
        items_raw = []
        for item in req.items:
            # item은 ProfileCreate 객체임
            raw = item.model_dump(by_alias=False)
            raw["metadata_"] = raw.pop("metadata_", None)
            
            if isinstance(raw.get("start_date"), str):
                raw["start_date"] = _parse_date(raw["start_date"])
            if isinstance(raw.get("end_date"), str):
                raw["end_date"] = _parse_date(raw["end_date"])
                
            items_raw.append(raw)
        
        profiles = await self.repo.bulk_create(user_id, items_raw)
        return [_to_response(p) for p in profiles]

    # ─── Unified Ingest ─────────────────────────────────────────

    async def ingest(
        self,
        user_id: str,
        source_type: str,
        enrichment_level: str = "basic",
        file: Optional[UploadFile] = None,
        text: Optional[str] = None,
        url: Optional[str] = None,
    ) -> IngestResponse:
        """
        통합 업로드 파이프라인.
        1) 텍스트 추출  2) AI 분류 → 즉시 저장  3) 심층 해석(optional)
        """
        from app.models.user import User
        from app.core.config import settings

        u_id = uuid.UUID(user_id)
        query = select(User).where(User.id == u_id).with_for_update()
        result = await self.db.execute(query)
        user = result.scalar_one_or_none()
        if not user:
            raise HTTPException(status_code=404, detail="사용자를 찾을 수 없습니다.")

        # ── Step 1: 텍스트 추출 ──
        extracted_text = ""
        source = ProfileSource.MANUAL

        if source_type == "file" and file:
            extracted_text = await self._extract_text_from_upload(file)
            source = ProfileSource.FILE_UPLOAD
        elif source_type == "text" and text:
            extracted_text = text
            source = ProfileSource.FREE_TEXT
        elif source_type == "link" and url:
            extracted_text = await self._extract_text_from_url(url)
            source = ProfileSource.LINK
        else:
            raise HTTPException(status_code=400, detail="source_type에 맞는 입력(file/text/url)이 필요합니다.")

        if not extracted_text or len(extracted_text.strip()) < 10:
            raise HTTPException(status_code=400, detail="추출된 텍스트가 너무 짧습니다.")

        # ── Step 2: AI 분류 (무료 ingest 차감 또는 포인트 체크) ──
        if not user.is_admin:
            if user.free_ingests_remaining > 0:
                user.free_ingests_remaining -= 1
            elif user.point_balance < 5:
                raise HTTPException(
                    status_code=status.HTTP_402_PAYMENT_REQUIRED,
                    detail="포인트가 부족합니다. (AI 분류: 5P 필요)"
                )
            else:
                from app.services.point_service import PointService
                point_service = PointService(self.db)
                await point_service.deduct_points(u_id, 5, "profile_ingest")

        parsed = await self._ai_categorize(extracted_text, settings)

        # ── Step 3: 즉시 저장 ──
        items_raw = []
        for item in parsed:
            raw = item.model_dump()
            raw["metadata_"] = raw.pop("metadata", None)
            raw["source"] = source
            raw["enrichment_status"] = "none"
            if isinstance(raw.get("start_date"), str):
                raw["start_date"] = _parse_date(raw["start_date"])
            if isinstance(raw.get("end_date"), str):
                raw["end_date"] = _parse_date(raw["end_date"])
            items_raw.append(raw)

        profiles = await self.repo.bulk_create(user_id, items_raw)
        profile_responses = [_to_response(p) for p in profiles]

        # ── Step 4: AI 요약 생성 ──
        ai_summary = self._build_quick_summary(profiles, extracted_text)

        # ── Step 5: 심층 해석 (enrichment_level=full) ──
        enrichment_status = "complete"
        if enrichment_level == "full":
            if not user.is_admin:
                if user.point_balance < 15:
                    enrichment_status = "skipped_insufficient_points"
                else:
                    from app.services.point_service import PointService
                    point_service = PointService(self.db)
                    await point_service.deduct_points(u_id, 15, "profile_enrichment")
                    await self._run_deep_enrichment(profiles, extracted_text, settings)
                    enrichment_status = "complete"
            else:
                await self._run_deep_enrichment(profiles, extracted_text, settings)
                enrichment_status = "complete"

            # 다시 응답 생성 (enrichment 반영)
            refreshed = await self.repo.list_by_user(user_id)
            saved_ids = {str(p.id) for p in profiles}
            profile_responses = [_to_response(p) for p in refreshed if str(p.id) in saved_ids]

        await self.db.commit()

        return IngestResponse(
            profiles=profile_responses,
            ai_summary=ai_summary,
            enrichment_status=enrichment_status,
        )

    async def _extract_text_from_upload(self, file: UploadFile) -> str:
        """업로드 파일에서 텍스트를 추출합니다."""
        content = await file.read()
        filename = file.filename or "unknown"
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        content_type = file.content_type or ""

        if ext == "pdf" or "pdf" in content_type:
            return self._extract_pdf(content)
        elif ext in ("xlsx", "xls") or "spreadsheet" in content_type:
            return self._extract_excel(content)
        elif ext in ("docx", "doc") or "word" in content_type:
            return self._extract_word(content)
        elif ext in ("png", "jpg", "jpeg", "webp") or content_type.startswith("image/"):
            return await self._extract_ocr(content)
        else:
            # 텍스트 파일 시도
            try:
                return content.decode("utf-8")
            except UnicodeDecodeError:
                return content.decode("euc-kr", errors="replace")

    def _extract_pdf(self, content: bytes) -> str:
        try:
            import fitz
            doc = fitz.open(stream=content, filetype="pdf")
            text = "\n".join(page.get_text() for page in doc)
            doc.close()
            return text
        except ImportError:
            raise HTTPException(status_code=500, detail="PDF 처리 모듈이 설치되지 않았습니다.")

    def _extract_excel(self, content: bytes) -> str:
        import io
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        lines = []
        for sheet in wb:
            for row in sheet.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    lines.append("\t".join(cells))
        wb.close()
        return "\n".join(lines)

    def _extract_word(self, content: bytes) -> str:
        import io
        from docx import Document
        doc = Document(io.BytesIO(content))
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())

    async def _extract_ocr(self, content: bytes) -> str:
        try:
            from paddleocr import PaddleOCR
            import tempfile, os
            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as f:
                f.write(content)
                tmp_path = f.name
            try:
                ocr = PaddleOCR(use_angle_cls=True, lang="korean", show_log=False)
                result = ocr.ocr(tmp_path, cls=True)
                lines = []
                if result and result[0]:
                    for line in result[0]:
                        lines.append(line[1][0])
                return "\n".join(lines)
            finally:
                os.unlink(tmp_path)
        except ImportError:
            raise HTTPException(status_code=500, detail="OCR 모듈이 설치되지 않았습니다.")

    async def _extract_text_from_url(self, url: str) -> str:
        """URL에서 텍스트를 추출합니다."""
        import httpx
        from bs4 import BeautifulSoup
        try:
            async with httpx.AsyncClient(timeout=10.0, follow_redirects=True) as client:
                resp = await client.get(url, headers={
                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
                })
                resp.raise_for_status()
                soup = BeautifulSoup(resp.text, "html.parser")
                for s in soup(["script", "style"]):
                    s.decompose()
                return soup.get_text(separator="\n", strip=True)[:8000]
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"링크에서 텍스트를 추출할 수 없습니다: {str(e)}")

    async def _ai_categorize(self, text: str, settings) -> list[ParsedProfileItem]:
        """GPT-4o-mini로 텍스트를 프로필 항목으로 분류합니다."""
        if not settings.OPENAI_API_KEY:
            return [ParsedProfileItem(
                profile_type="free_text",
                title="자유 텍스트 입력",
                description=text[:2000],
            )]

        from openai import AsyncOpenAI
        client = AsyncOpenAI(api_key=settings.OPENAI_API_KEY)

        prompt = f"""아래 텍스트에서 이력/경력 관련 정보를 추출하여 JSON 객체로 반환하세요.
반드시 "items"라는 키 아래에 배열을 담아서 반환해야 합니다.

각 항목은 다음 필드를 포함합니다:
- profile_type: education|work_experience|project|certification|skill|activity|award|strength|weakness|motivation|free_text
- title: 항목 제목 (필수)
- organization: 학교/회사/단체명
- role: 역할/직책
- description: 상세 내용 (STAR 기법으로 간결하게)
- start_date: YYYY-MM 형식
- end_date: YYYY-MM 형식
- tags: 관련 키워드 배열 (최소 3개)
- metadata: 추가 정보

텍스트:
{text[:8000]}

JSON 형식의 객체만 반환하세요. {{ "items": [{{...}}] }}"""

        resp = await client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": prompt}],
            response_format={"type": "json_object"},
            temperature=0.2,
        )
        data = json.loads(resp.choices[0].message.content)
        return [ParsedProfileItem(**item) for item in data.get("items", [])]

    def _build_quick_summary(self, profiles: list, extracted_text: str) -> AISummary:
        """저장된 프로필로부터 빠른 요약을 생성합니다 (AI 호출 없이)."""
        timeline = []
        all_tags = []
        strengths = []

        for p in profiles:
            timeline.append(TimelineEntry(
                title=p.title,
                organization=p.organization,
                start_date=str(p.start_date) if p.start_date else None,
                end_date=str(p.end_date) if p.end_date else None,
                profile_type=p.profile_type.value if hasattr(p.profile_type, 'value') else str(p.profile_type),
            ))
            if p.tags:
                all_tags.extend(p.tags)

        # 타입별 활용 제안
        type_suggestions = {
            "work_experience": "직무 역량 및 성장과정에 활용 가능",
            "project": "프로젝트 성과 및 문제해결력에 활용 가능",
            "education": "학업 역량 및 전공 적합성에 활용 가능",
            "certification": "전문성 증명에 활용 가능",
            "activity": "리더십/팀워크 역량에 활용 가능",
            "award": "성취 증명 및 지원동기에 활용 가능",
        }
        suggested_uses = []
        for p in profiles:
            pt = p.profile_type.value if hasattr(p.profile_type, 'value') else str(p.profile_type)
            if pt in type_suggestions and type_suggestions[pt] not in suggested_uses:
                suggested_uses.append(type_suggestions[pt])

        # 태그에서 스킬 추출
        skill_tags = list(set(all_tags))[:20]

        return AISummary(
            key_strengths=strengths,
            experience_timeline=timeline,
            skill_tags=skill_tags,
            suggested_uses=suggested_uses,
        )

    async def _run_deep_enrichment(self, profiles: list, extracted_text: str, settings) -> None:
        """GPT-4o로 각 프로필 항목을 심층 해석합니다."""
        if not settings.OPENAI_API_KEY:
            return

        from openai import AsyncOpenAI
        client = AsyncOpenAI(api_key=settings.OPENAI_API_KEY)

        for profile in profiles:
            desc = getattr(profile, "_description_plain", "") or ""
            if not desc:
                continue

            prompt = f"""당신은 전문 취업 컨설턴트입니다. 아래 경험 항목을 자소서 작성에 최적화된 전략 메모리로 변환하세요.

STAR 기법(Situation, Task, Action, Result)으로 재구성하고, 수치화된 성과를 포함하세요.
이 경험을 자소서의 어떤 항목(지원동기/성장과정/프로젝트 성과/직무 역량)에 활용하면 좋을지 제안하세요.

항목: {profile.title}
조직: {profile.organization or '미상'}
역할: {profile.role or '미상'}
내용: {desc[:3000]}

JSON 형식으로 반환:
{{ "interpreted_content": "STAR 기법으로 재구성한 내용", "suggested_section": "지원동기|성장과정|프로젝트 성과|직무 역량", "key_strength": "이 경험이 증명하는 핵심 역량" }}"""

            try:
                resp = await client.chat.completions.create(
                    model="gpt-4o",
                    messages=[{"role": "user", "content": prompt}],
                    response_format={"type": "json_object"},
                    temperature=0.3,
                )
                enrichment = json.loads(resp.choices[0].message.content)

                profile.ai_interpreted_content_encrypted = None
                from app.core.encryption import encrypt
                interpreted = enrichment.get("interpreted_content", "")
                if interpreted:
                    profile.ai_interpreted_content_encrypted = encrypt(interpreted)
                    profile._ai_interpreted_content_plain = interpreted

                profile.is_ai_memory = True
                profile.enrichment_status = "complete"
                profile.ai_summary_json = {
                    "suggested_section": enrichment.get("suggested_section"),
                    "key_strength": enrichment.get("key_strength"),
                }
            except Exception as e:
                logger.warning(f"Enrichment failed for profile {profile.id}: {e}")
                profile.enrichment_status = "none"

    # ─── Dashboard ────────────────────────────────────────────

    async def get_dashboard(self, user_id: str) -> DashboardResponse:
        """프로필 대시보드 데이터를 집계합니다."""
        profiles = await self.repo.list_by_user(user_id)

        # 타입별 카운트
        type_counts: dict[str, int] = {}
        for p in profiles:
            pt = p.profile_type.value if hasattr(p.profile_type, 'value') else str(p.profile_type)
            type_counts[pt] = type_counts.get(pt, 0) + 1

        # 완성도 점수: 핵심 타입 6개 중 몇 개를 채웠는가
        core_types = {"education", "work_experience", "project", "skill", "activity", "motivation"}
        filled = sum(1 for t in core_types if t in type_counts)
        completeness_score = int((filled / len(core_types)) * 100)

        # 타임라인
        timeline = []
        for p in profiles:
            if p.start_date or p.end_date:
                timeline.append(TimelineEntry(
                    title=p.title,
                    organization=p.organization,
                    start_date=str(p.start_date) if p.start_date else None,
                    end_date=str(p.end_date) if p.end_date else None,
                    profile_type=p.profile_type.value if hasattr(p.profile_type, 'value') else str(p.profile_type),
                ))
        timeline.sort(key=lambda t: t.start_date or "0000-00", reverse=True)

        # 태그 집계
        all_tags: list[str] = []
        for p in profiles:
            if p.tags:
                all_tags.extend(p.tags)
        tag_counter = Counter(all_tags)
        top_tags = [TagCount(tag=tag, count=count) for tag, count in tag_counter.most_common(15)]

        # 스킬 태그 (skill 타입 프로필 기반)
        skill_tags = [TagCount(tag=tag, count=count) for tag, count in tag_counter.most_common(10)]

        # AI 강점 요약 (프로필 3개 이상이면 생성)
        ai_strength_summary = None
        if len(profiles) >= 3:
            ai_strength_summary = await self._generate_strength_summary(profiles)

        return DashboardResponse(
            total_profiles=len(profiles),
            type_counts=type_counts,
            completeness_score=completeness_score,
            timeline=timeline,
            skill_tags=skill_tags,
            ai_strength_summary=ai_strength_summary,
            top_tags=top_tags,
        )

    async def _generate_strength_summary(self, profiles: list) -> str | None:
        """GPT-4o-mini로 전체 프로필 기반 강점 요약을 생성합니다."""
        from app.core.config import settings
        if not settings.OPENAI_API_KEY:
            return None

        try:
            from openai import AsyncOpenAI
            client = AsyncOpenAI(api_key=settings.OPENAI_API_KEY)

            profile_text = "\n".join([
                f"- [{p.profile_type.value}] {p.title} ({p.organization or ''}) {p.role or ''}"
                for p in profiles[:15]
            ])

            resp = await client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[{"role": "user", "content": f"""아래 이력 항목들을 분석하여, 이 지원자의 핵심 강점 3가지를 한국어로 간결하게 요약하세요.
각 강점은 1-2문장으로, 구체적인 근거와 함께 설명하세요.

이력:
{profile_text}

형식: "1. 강점: 설명\n2. 강점: 설명\n3. 강점: 설명" """}],
                temperature=0.3,
                max_tokens=500,
            )
            return resp.choices[0].message.content
        except Exception as e:
            logger.warning(f"Strength summary generation failed: {e}")
            return None

    async def interpret_file_to_memory(self, user_id: str, filename: str, text: str) -> ProfileResponse:
        """파일 텍스트를 AI가 자소서용 전략 메모리로 해석하여 즉시 저장합니다."""
        from app.models.user import User
        from sqlalchemy import select
        
        u_id = uuid.UUID(user_id)
        
        # 1. 포인트 체크 (30P 차감 또는 어드민 무제한)
        from app.services.point_service import PointService
        point_service = PointService(self.db)
        await point_service.deduct_points(
            user_id=u_id,
            amount=30,
            reason=f"AI Experience Interpretation: {filename}"
        )

        from app.core.config import settings
        from openai import AsyncOpenAI
        client = AsyncOpenAI(api_key=settings.OPENAI_API_KEY)

        # 2. AI 해석 프롬프트
        # 자소서 작성에 최적화된 형식으로 텍스트를 재구성
        prompt = f"""당신은 전문 취업 컨설턴트입니다. 아래 텍스트는 사용자가 업로드한 파일(포트폴리오, 프로젝트, 경력기술서 등)의 내용입니다.
이 내용을 바탕으로 나중에 '자기소개서 작성 프롬프트'에서 AI가 참조할 수 있는 '핵심 경험 메모리'로 변환하세요.

변환 지침:
1. 단순 텍스트 추출이 아니라, "무엇을 했고(Action)", "어떤 성과를 냈으며(Result)", "이를 통해 어떤 역량을 증명(Competency)"할 수 있는지 중심으로 재구성하세요.
2. STAR 기법(Situation, Task, Action, Result)으로 정리할 수 있다면 그렇게 하세요.
3. 수치화된 성과가 있다면 반드시 포함하세요.
4. 이 내용을 자소서의 '지원동기', '성장과정', '프로젝트 성과' 중 어디에 쓰면 좋을지도 짧게 제안하세요.

입력 텍스트:
{text[:10000]}  # 컨텍스트 제한 고려

반환 형식:
반드시 다음 JSON 구조로 응답하세요:
{{
  "title": "경험의 대표 제목",
  "organization": "관련 조직명",
  "role": "수행 직무/역할",
  "interpreted_content": "AI가 재구성한 자소서용 핵심 메모리 내용 전체",
  "tags": ["키워드1", "키워드2"],
  "profile_type": "project|work_experience|activity 중 가장 적절한 것"
}}
"""
        resp = await client.chat.completions.create(
            model="gpt-4o",
            messages=[{"role": "user", "content": prompt}],
            response_format={"type": "json_object"},
            temperature=0.3,
        )
        
        interpreted = json.loads(resp.choices[0].message.content)
        
        # 3. DB 저장
        profile_data = {
            "profile_type": interpreted.get("profile_type", "project"),
            "title": interpreted.get("title", f"AI 해석: {filename}"),
            "organization": interpreted.get("organization"),
            "role": interpreted.get("role"),
            "description": text[:2000],  # 원본 텍스트 일부 보존
            "is_ai_memory": True,
            "ai_interpreted_content": interpreted.get("interpreted_content"),
            "tags": interpreted.get("tags", []),
            "source": ProfileSource.FILE_UPLOAD
        }
        
        p = await self.repo.create(user_id, profile_data)
        return _to_response(p)
